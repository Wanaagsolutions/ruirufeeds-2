import base64
import logging
import os
import tempfile
import openpyxl as xl

from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from odoo import models, fields, api
from ..models.stock import EXPENSE_TYPES

BIG_FONT = Font(name='Arial', bold=True, size=15)
MID_FONT = Font(name='Arial', bold=True, size=12)
NORMAL_FONT = Font(name='Arial', size=12)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')


_logger = logging.getLogger(__name__)

TRIP_ANALYSIS_COL = [
    'Truck', 'POD No', 'Trip', 'Driver Name', 'Customer', 'Start Date', 'End Date',
    'Loaded QTY OBS', 'Loaded QTY @20',	'Offloaded QTY OBS', 'Offloaded QTY@ 20', 'Shortage',
    'Allowable Short', 'Abnormal Short', 'Pickup Location', 'Drop off Location', 'Shortage Rate', 'Cargo',
    'Currency',	'AR Rate', 'Revenue'
]
TRIP_ANALYSIS_COL.extend([f'{i[1]} Expense' for i in EXPENSE_TYPES])
TRIP_ANALYSIS_COL.append('Margin')

TRUCK_ANALYSIS_COL = [
    'Truck', 'Driver Name', 'Loaded QTY OBS', 'Loaded QTY @20',	'Offloaded QTY OBS', 'Offloaded QTY@ 20', 'Shortage',
    'Allowable Short', 'Abnormal Short', 'Cargo', 'Revenue'
]
TRUCK_ANALYSIS_COL.extend([f'{i[1]} Expense' for i in EXPENSE_TYPES])
TRUCK_ANALYSIS_COL.append('Margin')


class FleetAnalysis(models.TransientModel):
    _name = 'fleet.analysis'
    _description = 'Fleet Margin Analysis'

    date_from = fields.Date(string='Date From', required=True)
    date_to = fields.Date(string='Date To', required=True, default=fields.Date.today())
    report_type = fields.Selection(string='Report Type',
                                   selection=[('trip', 'Trip'), ('truck', 'Truck')], default='trip')
    
    def _trips_domain(self):
        return [
            ('offloading_date', '>=', self.date_from), 
            ('offloading_date', '<=', self.date_to), 
            ('state', '!=', 'cancelled')
            ]
    
    def _expense_lines(self, trips):
        analytics = trips.mapped('analytic_tag_id').ids
        analytic_accounts = trips.mapped('truck_id.analytic_account_id').ids
        domain = [
                ('analytic_tag_ids', 'in', analytics),
                ('analytic_account_id', 'in', analytic_accounts),
                ('move_id.state', '=', 'posted'),
                ('product_id.can_be_expensed', '=', True)
            ]
        return self.env['account.move.line'].search(domain)
              
    def _prepare_trip_analyis_data(self):
        values = []
        trips = self.env['fleet.trip'].search(self._trips_domain(), order='name')
        for trip in trips:
            line_ids = self._expense_lines(trip)

            vals = {
                'Truck': trip.truck_id.name,
                'Driver Name': trip.driver_id.name,
                'Customer': trip.partner_id.name,
                'Start Date': trip.date,
                'End Date': trip.end_date,
                'Loaded QTY OBS': trip.loaded_qty,
                'Loaded QTY @20': trip.loaded_qty_20,
                'Offloaded QTY OBS': trip.offloaded_qty,
                'Offloaded QTY@ 20': trip.offloaded_qty_20,
                'Allowable Short': trip.allowable_shortage,
                'Abnormal Short': trip.short_qty,
                'Pickup Location': trip.pickup_location,
                'Drop off Location': trip.drop_location,
                'Shortage Rate': trip.shortage_rate,
                'Shortage': trip.shortage_rate * trip.short_qty,
                'Cargo': trip.product_id.name,
                'Currency': trip.currency_id.name,
                'AR Rate': trip.order_id.ar_rate,
                'Revenue': trip.loaded_qty_20 * trip.order_id.ar_rate,
                'Margin': (trip.loaded_qty_20 * trip.order_id.ar_rate),
            }
            for expense in EXPENSE_TYPES:
                expense_lines = line_ids.filtered(lambda l: l.product_id.expense_type == expense[0])
                expense_val = sum(
                    map(
                        lambda e: e.company_id.currency_id._convert(abs(e.debit), trip.currency_id, trip.company_id, e.date), 
                        expense_lines
                        )
                    )
                vals[f'{expense[1]} Expense'] = expense_val
                vals['Margin'] -= expense_val
            
            values.append(vals)
        return values
    
    def _prepare_truck_analysis_data(self):
        values = defaultdict(dict)
        trips = self.env['fleet.trip'].search(self._trips_domain(), order='truck_id')
        
        for trip in trips:
            line_ids = self._expense_lines(trip)
            vals = {
                'Truck': trip.truck_id.name,
                'Driver Name': trip.truck_id.driver_id.name,
                'Loaded QTY OBS': trip.loaded_qty,
                'Loaded QTY @20': trip.loaded_qty_20,
                'Offloaded QTY OBS': trip.offloaded_qty,
                'Offloaded QTY@ 20': trip.offloaded_qty_20,
                'Allowable Short': trip.allowable_shortage,
                'Abnormal Short': trip.short_qty,
                'Shortage': trip.shortage_rate * trip.short_qty,
                'Cargo': trip.product_id.name,
                'Revenue': trip.loaded_qty_20 * trip.order_id.ar_rate,
                'Margin': (trip.loaded_qty_20 * trip.order_id.ar_rate),
            }

            for expense in EXPENSE_TYPES:
                expense_lines = line_ids.filtered(lambda l: l.product_id.expense_type == expense[0])
                expense_val = sum(
                    map(
                        lambda e: e.company_id.currency_id._convert(abs(e.debit), trip.currency_id, trip.company_id, e.date), 
                        expense_lines
                        )
                    )
                vals[f'{expense[1]} Expense'] = expense_val
                vals['Margin'] -= expense_val
                
            if values[trip.truck_id]:
                for col in set(TRUCK_ANALYSIS_COL) - set(['Truck', 'Driver Name', 'Cargo']):
                    values[trip.truck_id][col] += vals.get(col)
            else:
                values[trip.truck_id] = vals
        return list(values.values())
    
    def _add_header_row(self, sheet, cols, row=1, start_col=1, font=MID_FONT):
        for index, col in enumerate(cols, start=start_col):
            letter = get_column_letter(index)
            sheet[f'{letter}{row}'] = col
            sheet[f'{letter}{row}'].font = font
    
    def _add_value_row(self, sheet, data, cols, row=1, start_col=1, font=NORMAL_FONT):
        for index, col in enumerate(cols, start=start_col):
            letter = get_column_letter(index)
            sheet[f'{letter}{row}'] = data.get(col)
            sheet[f'{letter}{row}'].font = font

    def _generate_trips_report(self, wb):
        ws = wb.active
        fr = 1
        self._add_header_row(ws, TRIP_ANALYSIS_COL)
        fr += 1
        for val in self._prepare_trip_analyis_data():
            self._add_value_row(ws, val, TRIP_ANALYSIS_COL, fr)
            fr += 1
        return wb, 'Trips Analysis'

    def _generate_trucks_report(self, wb):
        ws = wb.active
        fr = 1
        self._add_header_row(ws, TRUCK_ANALYSIS_COL)
        fr += 1
        
        for val in sorted(self._prepare_truck_analysis_data(), key=lambda k: k['Truck']):
            for index, col in enumerate(TRUCK_ANALYSIS_COL, start=1):
                letter = get_column_letter(index)
                ws[f'{letter}{fr}'] = val.get(col)
                ws[f'{letter}{fr}'].font = NORMAL_FONT

            fr += 1
        return wb, 'Truck Analysis'
    
    def action_generate_report(self):
        wb = xl.Workbook()
        if self.report_type == 'trip':
            wb, filename = self._generate_trips_report(wb)
        else:
            wb, filename = self._generate_trucks_report(wb)
            
        xlsx_path = self.env['excel.wizard'].create_xls()
        wb.save(xlsx_path)
        mes = f'{filename} generated by {self.env.user.name}.'
        datas = self.env['excel.wizard'].save_xls_file(xlsx_path, mes)

        attachment_id = self.env['excel.wizard'].create({
            'name': f'{filename}.xlsx',
            'report': datas
        })
        return {
            'name': filename,
            'context': self.env.context,
            'view_mode': 'form',
            'res_model': 'excel.wizard',
            'res_id': attachment_id.id,
            'type': 'ir.actions.act_window',
            'target': 'new'
        }

class ReportDownload(models.TransientModel):
    _name = 'excel.wizard'
    _description = 'Download Excel Forms'

    name = fields.Char('File Name', size=64)
    report = fields.Binary('Your Report', readonly=True)

    @api.model
    def create_xls(self):
        _, xls_path = tempfile.mkstemp(
            suffix='.xlsx', prefix='xlsreport.tmp.')
        return xls_path

    @api.model
    def save_xls_file(self, xls_path, message=None):
        if not message:
            message = f'A report has been generated by {self.env.user.name}'
        with open(xls_path, 'rb') as f:
            datas = base64.encodebytes(f.read())
            _logger.info(message)
            self.delete_tempfile(xls_path)
        return datas

    @api.model
    def delete_tempfile(self, path):
        try:
            os.unlink(path)
        except (OSError, IOError):
            _logger.error('Error when trying to remove file %s' % path)
